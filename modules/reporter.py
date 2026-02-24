"""Модуль формирования Excel-отчётов.

Генерирует Excel-реестр документов с метаданными, автофильтрами,
форматированием и отдельным листом для файлов, требующих проверки.
Включает лист "Сводка" с диаграммами и аудит-блоком.
"""
import logging
from collections import Counter
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Маппинг ключей БД → заголовки Excel (порядок сохраняется)
COLUMNS = {
    "filename": "Файл",
    "contract_type": "Тип документа",
    "counterparty": "Контрагент",
    "subject": "Предмет",
    "date_signed": "Дата подписания",
    "date_start": "Начало",
    "date_end": "Окончание",
    "amount": "Сумма",
    "special_conditions": "Особые условия",
    "confidence": "Уверенность AI",
    "validation_status": "Статус валидации",
    "validation_warnings": "Замечания",
    "status": "Статус обработки",
    "review_status": "Статус проверки",
    "lawyer_comment": "Комментарий юриста",
}

_REVIEW_LABELS = {
    "not_reviewed": "Не проверен",
    "reviewed": "Проверен",
    "attention_needed": "Требует внимания",
}

STATUS_COLORS = {
    "ok": "C6EFCE",        # Зелёный
    "warning": "FFEB9C",   # Жёлтый
    "unreliable": "FFC7CE",  # Красный
    "error": "FFC7CE",     # Красный
}


def generate_report(db_results: list[dict], output_dir: Path) -> Path:
    """Генерирует Excel-файл с двумя листами.

    Лист 1: «Реестр договоров» — все файлы с метаданными.
    Лист 2: «Требуют проверки» — только файлы с validation_status != "ok".

    Возвращает путь к созданному файлу.
    """
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / "Реестр_договоров.xlsx"

    # Подготовить данные
    df = _prepare_dataframe(db_results)

    # Лист с проблемными файлами
    df_problems = df[
        df.get("Статус валидации", pd.Series(dtype=str)).isin(
            ["warning", "unreliable", "error"]
        )
    ]

    # Записать в Excel
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Реестр договоров", index=False)
        if not df_problems.empty:
            df_problems.to_excel(writer, sheet_name="Требуют проверки", index=False)

    # Форматирование через openpyxl
    wb = load_workbook(output_path)
    _format_sheet(wb["Реестр договоров"])
    if "Требуют проверки" in wb.sheetnames:
        _format_sheet(wb["Требуют проверки"])

    # Добавить комментарии с замечаниями валидации
    _add_validation_comments(wb["Реестр договоров"])

    # Создать лист Сводка (первый по порядку)
    _create_summary_sheet(wb, db_results)

    wb.save(output_path)

    logger.info("Excel-реестр создан: %s (%d записей)", output_path, len(df))
    return output_path


def _prepare_dataframe(db_results: list[dict]) -> pd.DataFrame:
    """Готовит DataFrame: переименовывает колонки, конвертирует списки в строки."""
    df = pd.DataFrame(db_results)

    # Конвертировать списки в читаемые строки (до переименования)
    for col in ("special_conditions", "validation_warnings", "parties"):
        if col in df.columns:
            df[col] = df[col].apply(
                lambda v: "; ".join(v) if isinstance(v, list) else (v or "")
            )

    # Маппинг review_status → читаемые метки (v0.3)
    if "review_status" in df.columns:
        df["review_status"] = df["review_status"].map(_REVIEW_LABELS).fillna("Не проверен")

    # Переименовать колонки
    df = df.rename(columns=COLUMNS)

    # Оставить только нужные столбцы в правильном порядке
    cols = [v for v in COLUMNS.values() if v in df.columns]
    return df[cols]


def _format_sheet(ws) -> None:
    """Применяет форматирование к листу Excel."""
    # Заголовки: жирный шрифт, синий фон
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    # Автофильтр
    ws.auto_filter.ref = ws.dimensions

    # Фиксация первой строки
    ws.freeze_panes = "A2"

    # Цветовая подсветка по статусу валидации
    status_col = None
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == "Статус валидации":
            status_col = col_idx
            break

    if status_col:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            status_cell = row[status_col - 1]
            if status_cell.value in STATUS_COLORS:
                fill = PatternFill(
                    start_color=STATUS_COLORS[status_cell.value],
                    end_color=STATUS_COLORS[status_cell.value],
                    fill_type="solid",
                )
                for cell in row:
                    cell.fill = fill

    # Автоширина столбцов (выборка до 50 строк для производительности)
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(
            min_row=1, max_row=min(ws.max_row, 50),
            min_col=col_idx, max_col=col_idx,
        ):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 50)


def _add_validation_comments(ws) -> None:
    """Добавляет комментарии к ячейкам статуса с текстом замечаний."""
    status_col = None
    warnings_col = None
    for col_idx, cell in enumerate(ws[1], 1):
        if cell.value == "Статус валидации":
            status_col = col_idx
        elif cell.value == "Замечания":
            warnings_col = col_idx

    if not status_col or not warnings_col:
        return

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        warnings_cell = row[warnings_col - 1]
        status_cell = row[status_col - 1]
        if warnings_cell.value and status_cell.value in ("warning", "unreliable", "error"):
            status_cell.comment = Comment(
                str(warnings_cell.value), "ЮрТэг", width=350, height=100,
            )


def _create_summary_sheet(wb, db_results: list[dict]) -> None:
    """Создаёт лист Сводка с диаграммами и аудит-блоком."""
    ws = wb.create_sheet("Сводка", 0)  # Первый лист

    title_font = Font(bold=True, size=14, color="4472C4")
    header_font = Font(bold=True, size=11)
    label_font = Font(color="64748B", size=10)

    # Заголовок
    ws["A1"] = "Сводка по обработке договоров"
    ws["A1"].font = title_font
    ws.merge_cells("A1:D1")

    # Общие метрики
    total = len(db_results)
    done = sum(1 for r in db_results if r.get("status") == "done")
    errors = sum(1 for r in db_results if r.get("status") == "error")
    ok_count = sum(1 for r in db_results if r.get("validation_status") == "ok")
    warn_count = sum(
        1 for r in db_results
        if r.get("validation_status") in ("warning", "unreliable", "error")
    )

    ws["A3"] = "Всего файлов"
    ws["A3"].font = label_font
    ws["B3"] = total
    ws["B3"].font = header_font

    ws["A4"] = "Успешно обработано"
    ws["A4"].font = label_font
    ws["B4"] = done
    ws["B4"].font = header_font

    ws["A5"] = "Ошибки обработки"
    ws["A5"].font = label_font
    ws["B5"] = errors
    ws["B5"].font = header_font

    ws["A6"] = "Статус OK"
    ws["A6"].font = label_font
    ws["B6"] = ok_count
    ws["B6"].font = header_font

    ws["A7"] = "Требуют проверки"
    ws["A7"].font = label_font
    ws["B7"] = warn_count
    ws["B7"].font = header_font

    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 15

    # --- Данные для Pie Chart (типы договоров) ---
    type_counts = Counter(
        r.get("contract_type", "Не определён") or "Не определён"
        for r in db_results if r.get("status") == "done"
    )

    # Записываем данные в ячейки (скрытая таблица справа)
    ws["F1"] = "Тип договора"
    ws["F1"].font = header_font
    ws["G1"] = "Количество"
    ws["G1"].font = header_font
    for i, (ctype, count) in enumerate(type_counts.most_common(10), start=2):
        ws[f"F{i}"] = ctype
        ws[f"G{i}"] = count

    if type_counts:
        pie = PieChart()
        pie.title = "Типы договоров"
        pie.style = 10
        data_ref = Reference(ws, min_col=7, min_row=1, max_row=1 + len(type_counts))
        cats_ref = Reference(ws, min_col=6, min_row=2, max_row=1 + len(type_counts))
        pie.add_data(data_ref, titles_from_data=True)
        pie.set_categories(cats_ref)
        pie.width = 16
        pie.height = 12
        ws.add_chart(pie, "A9")

    # --- Данные для Bar Chart (статусы валидации) ---
    status_counts = Counter(
        r.get("validation_status", "—") or "—"
        for r in db_results if r.get("status") == "done"
    )

    ws["F15"] = "Статус"
    ws["F15"].font = header_font
    ws["G15"] = "Количество"
    ws["G15"].font = header_font
    for i, (status, count) in enumerate(status_counts.most_common(), start=16):
        ws[f"F{i}"] = status
        ws[f"G{i}"] = count

    if status_counts:
        bar = BarChart()
        bar.title = "Статусы валидации"
        bar.style = 10
        bar.type = "col"
        data_ref = Reference(ws, min_col=7, min_row=15, max_row=15 + len(status_counts))
        cats_ref = Reference(ws, min_col=6, min_row=16, max_row=15 + len(status_counts))
        bar.add_data(data_ref, titles_from_data=True)
        bar.set_categories(cats_ref)
        bar.width = 16
        bar.height = 12
        # Цвета
        from openpyxl.chart.series import DataPoint
        from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
        ws.add_chart(bar, "A25")

    # --- Аудит-блок ---
    audit_row = 42
    ws[f"A{audit_row}"] = "Параметры обработки"
    ws[f"A{audit_row}"].font = Font(bold=True, size=11, color="64748B")

    ws[f"A{audit_row + 1}"] = "Дата обработки:"
    ws[f"A{audit_row + 1}"].font = label_font
    ws[f"B{audit_row + 1}"] = datetime.now().strftime("%Y-%m-%d %H:%M")

    ws[f"A{audit_row + 2}"] = "Модель AI:"
    ws[f"A{audit_row + 2}"].font = label_font
    # Берём из первого результата
    model = "—"
    for r in db_results:
        if r.get("model_used"):
            model = r["model_used"]
            break
    ws[f"B{audit_row + 2}"] = model

    ws[f"A{audit_row + 3}"] = "Версия:"
    ws[f"A{audit_row + 3}"].font = label_font
    ws[f"B{audit_row + 3}"] = "ЮрТэг v0.3"

    # Уникальные контрагенты
    counterparties = set(
        r.get("counterparty", "") for r in db_results
        if r.get("counterparty") and r.get("status") == "done"
    )
    ws[f"A{audit_row + 4}"] = "Уникальных контрагентов:"
    ws[f"A{audit_row + 4}"].font = label_font
    ws[f"B{audit_row + 4}"] = len(counterparties)
