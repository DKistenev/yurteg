"""Тесты модуля reporter — генерация Excel-реестра."""
import pytest
from pathlib import Path

import openpyxl

from modules.reporter import generate_report, COLUMNS


def _make_contract(overrides: dict = None) -> dict:
    """Вспомогательная функция для создания тестовой записи договора."""
    base = {
        "filename": "test_contract.pdf",
        "contract_type": "Договор оказания услуг",
        "counterparty": "ООО Тест",
        "subject": "Оказание IT-услуг",
        "date_signed": "2024-01-15",
        "date_start": "2024-02-01",
        "date_end": "2025-01-31",
        "amount": "100 000 ₽",
        "special_conditions": ["Условие 1", "Условие 2"],
        "confidence": 0.9,
        "validation_status": "ok",
        "validation_warnings": [],
        "status": "done",
        "review_status": "not_reviewed",
        "lawyer_comment": None,
        "model_used": "test-model",
    }
    if overrides:
        base.update(overrides)
    return base


# ── generate_report ────────────────────────────────────────────────────────────

def test_generate_report_creates_file(tmp_path):
    """generate_report создаёт Excel-файл."""
    contracts = [_make_contract()]
    result_path = generate_report(contracts, tmp_path)
    assert result_path.exists()
    assert result_path.suffix == ".xlsx"


def test_generate_report_empty_data(tmp_path):
    """Пустой список создаёт Excel только с заголовками."""
    result_path = generate_report([], tmp_path)
    wb = openpyxl.load_workbook(result_path)
    # Лист реестра должен существовать
    assert "Реестр договоров" in wb.sheetnames
    ws = wb["Реестр договоров"]
    # Только строка с заголовками (max_row = 1 или 0)
    assert ws.max_row <= 1


def test_generate_report_single_contract(tmp_path):
    """Один договор → одна строка данных в листе реестра."""
    contracts = [_make_contract({"counterparty": "ООО Единственный"})]
    result_path = generate_report(contracts, tmp_path)
    wb = openpyxl.load_workbook(result_path)
    ws = wb["Реестр договоров"]
    # Строка 1 = заголовки, строка 2 = данные
    assert ws.max_row == 2
    # Значения строки 2 должны содержать наше имя контрагента
    row_values = [cell.value for cell in ws[2]]
    assert "ООО Единственный" in row_values


def test_generate_report_unicode(tmp_path):
    """Кирилличский контрагент с кавычками корректно записывается в Excel."""
    contracts = [_make_contract({"counterparty": "ООО «Ромашка»"})]
    result_path = generate_report(contracts, tmp_path)
    wb = openpyxl.load_workbook(result_path)
    ws = wb["Реестр договоров"]
    row_values = [cell.value for cell in ws[2]]
    assert "ООО «Ромашка»" in row_values


def test_generate_report_columns(tmp_path):
    """Заголовки листа реестра соответствуют значениям COLUMNS."""
    contracts = [_make_contract()]
    result_path = generate_report(contracts, tmp_path)
    wb = openpyxl.load_workbook(result_path)
    ws = wb["Реестр договоров"]
    header_values = [cell.value for cell in ws[1] if cell.value]
    expected_headers = list(COLUMNS.values())
    # Все ожидаемые заголовки должны присутствовать в листе
    for h in expected_headers:
        assert h in header_values, f"Заголовок '{h}' не найден в Excel"


def test_generate_report_review_sheet(tmp_path):
    """Договор со статусом warning попадает на лист 'Требуют проверки'."""
    contracts = [
        _make_contract({"validation_status": "ok"}),
        _make_contract({
            "filename": "problem.pdf",
            "validation_status": "warning",
            "counterparty": "ООО Проблемный",
        }),
    ]
    result_path = generate_report(contracts, tmp_path)
    wb = openpyxl.load_workbook(result_path)
    assert "Требуют проверки" in wb.sheetnames
    ws = wb["Требуют проверки"]
    # Должна быть одна строка данных
    assert ws.max_row == 2  # заголовок + 1 проблемный


def test_generate_report_summary_sheet(tmp_path):
    """Лист Сводка создаётся как первый лист."""
    contracts = [_make_contract()]
    result_path = generate_report(contracts, tmp_path)
    wb = openpyxl.load_workbook(result_path)
    assert "Сводка" in wb.sheetnames
    # Проверяем что в сводке есть заголовок
    ws = wb["Сводка"]
    assert ws["A1"].value is not None
